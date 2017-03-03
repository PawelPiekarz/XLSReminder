import openpyxl
import datetime
from datetime import timedelta
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText


sender_mail = ''
password = ''
recipient_mail = ''


def find_task(file):
    xls = openpyxl.load_workbook(filename=file)
    sheet = xls.get_sheet_by_name('Arkusz1')
    delta = 5
    date_min = datetime.datetime.now() + timedelta(days=-delta)
    date_max = datetime.datetime.now() + timedelta(days=delta)
    for row_index in range(1, sheet.max_row):
        if sheet.cell(row=row_index, column=2).value >= date_min and sheet.cell(row=row_index, column=2).value <= date_max:
            if not sheet.cell(row=row_index, column=3).value == 1:
                task = sheet.cell(row=row_index, column=1).value
                deadline = sheet.cell(row=row_index, column=2).value
                print(task)
                send_message(task, deadline)
                sheet.cell(row=row_index, column=3).value = 1
    try:
        xls.save(filename=file)
    except PermissionError:
        print("I can't save it, check if file is closed")


def send_message(task,deadline):
    msg = MIMEMultipart()
    msg['From'] = sender_mail
    msg['To'] = recipient_mail
    msg['Subject'] = 'Task reminder'
    message = """
    Hi!
    I found that your task: {} 
    is close to it's deadline:  {}

    Thank You!
    """.format(task, deadline.strftime("%d/%m/%y"))
    msg.attach(MIMEText(message))

    mailserver = smtplib.SMTP('smtp.gmail.com', 587)
    mailserver.ehlo()
    mailserver.starttls()
    mailserver.ehlo()
    mailserver.login(sender_mail, password)

    mailserver.sendmail(sender_mail, recipient_mail, msg.as_string())

    mailserver.quit()


if __name__ == "__main__":
    path = "test.xlsx"
    find_task(path)

